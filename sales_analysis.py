import pandas as pd
# --- Openpyxl Imports for Advanced Formatting ---
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment

# --- Configuration ---
SALES_FILE = "python_test-sales.csv"
PRODUCT_FILE = "python_test-product.csv"
STORE_FILE = "python_test-store.csv"
CURRENCY_FILE = "currency_rates.csv"
# Updated output file to include filter context (will be changed in main)
OUTPUT_EXCEL_FILE = "sales_reports_MYR_openpyxl.xlsx" 
REPORT_COLUMNS = ['sales_qty', 'sales_amount', 'sales_cost', 'profit']

def clean_currency_data(currency_df: pd.DataFrame) -> pd.DataFrame:
    """
    Cleans the currency exchange rate DataFrame by adjusting rates quoted per 100 units.
    
    If a column header ends with '100', it divides the rates by 100 
    and removes '100' from the column name (e.g., 'JPY100' -> 'JPY').
    """
    print("   - Cleaning currency data (adjusting rates quoted per 100 units)...")
    
    new_columns = {}
    for col in currency_df.columns:
        # Check if the column name ends with '100' (and is not the 'rate_date' column)
        if col.endswith('100'):
            # 1. Divide data by 100
            # Pandas automatically handles the division for the entire column
            currency_df[col] = currency_df[col] / 100
            
            # 2. Rename column (remove the '100' suffix)
            new_col_name = col[:-3]
            new_columns[col] = new_col_name
            print(f"      - Adjusted rate for {col}: new column {new_col_name}")
    
    # Apply all column renames at once
    currency_df = currency_df.rename(columns=new_columns)
    
    return currency_df


def load_data():
    """
    1. Reads all four sales, product, store, and currency CSV files.
    """
    print("1. Reading data from CSV files...")
    try:
        sales_df = pd.read_csv(SALES_FILE)
        product_df = pd.read_csv(PRODUCT_FILE)
        store_df = pd.read_csv(STORE_FILE)
        
        # Load currency data and rename date column
        currency_df = pd.read_csv(CURRENCY_FILE, index_col=False)
        currency_df = currency_df.rename(columns={currency_df.columns[0]: 'rate_date'})
        
        # Clean currency data immediately after loading
        currency_df = clean_currency_data(currency_df)

        print(f"   - Loaded {len(sales_df)} sales records.")
        return sales_df, product_df, store_df, currency_df
    except FileNotFoundError as e:
        print(f"Error: File not found: {e}")
        raise

def process_sales_data(sales_df, product_df, store_df, currency_df, store_region_filter=None, product_category_filter=None):
    """
    2. Merges data, converts prices/costs to MYR, calculates profit, and filters.
    """
    print("\n2. Processing and converting data to MYR...")
    
    # --- Data Merging ---
    merged_df = pd.merge(sales_df, product_df, on='product_code', how='left')
    merged_df = pd.merge(merged_df, store_df, on='store_code', how='left')

    # --- Currency Conversion (Using the first exchange rate row) ---
    fixed_rate_row = currency_df.iloc[0]
    exchange_rates = {'MYR': 1.0}
    
    # Build dictionary of exchange rates (e.g., {'USD': 4.15, 'SGD': 3.16})
    for currency in merged_df['currency'].unique():
        # This now correctly looks for cleaned currency codes like 'JPY'
        if currency != 'MYR' and currency in fixed_rate_row:
            exchange_rates[currency] = fixed_rate_row[currency]
        elif currency not in exchange_rates:
            print(f"Warning: No rate found for {currency}. Assuming rate of 1.0.")
            exchange_rates[currency] = 1.0

    merged_df['exchange_rate'] = merged_df['currency'].map(exchange_rates)
    
    # --- Calculations (Converted to MYR) ---
    # sales_amount = sum of (sales_qty * price) * rate
    merged_df['sales_amount_myr'] = merged_df['sales_qty'] * merged_df['price'] * merged_df['exchange_rate']
    # sales_cost = sum of (sales_qty * cost) * rate
    merged_df['sales_cost_myr'] = merged_df['sales_qty'] * merged_df['cost'] * merged_df['exchange_rate']
    # profit = sales_amount - sales_cost
    merged_df['profit_myr'] = merged_df['sales_amount_myr'] - merged_df['sales_cost_myr']

    # --- Filtering ---
    filtered_df = merged_df.copy()
    initial_count = len(filtered_df)
    
    if store_region_filter:
        print(f"   - Filtering by Store Region: {store_region_filter}")
        filtered_df = filtered_df[filtered_df['store_region'] == store_region_filter]
    if product_category_filter:
        print(f"   - Filtering by Product Category: {product_category_filter}")
        filtered_df = filtered_df[filtered_df['product_category'] == product_category_filter]
        
    print(f"   - Total records after filtering: {len(filtered_df)} (from {initial_count}).")
    
    return filtered_df

def aggregate_reports(df):
    """
    3. Groups data to create reports by region and category.
    """
    print("\n3. Aggregating sales data...")
    
    agg_cols = {
        'sales_qty': 'sum',
        'sales_amount_myr': 'sum',
        'sales_cost_myr': 'sum',
        'profit_myr': 'sum'
    }

    # Helper function for grouping and renaming columns
    def create_report(group_col):
        report = df.groupby(group_col).agg(agg_cols).reset_index()
        report = report.rename(columns={
            'sales_amount_myr': 'sales_amount',
            'sales_cost_myr': 'sales_cost',
            'profit_myr': 'profit'
        })
        report['sales_qty'] = report['sales_qty'].astype(int)
        return report

    report_region = create_report('store_region')
    report_category = create_report('product_category')
    
    print("   - Generated reports by Region and Category.")
    return report_region, report_category

def export_to_excel(report_region, report_category, filename):
    """
    4. Exports the reports to an Excel file using the 'openpyxl' engine 
       and applies formatting using openpyxl objects.
    """
    print(f"\n4. Exporting data to Excel file: '{filename}' using openpyxl...")
    
    # Define styles for openpyxl
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Use pandas ExcelWriter with the openpyxl engine
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # Write reports to different sheets
        report_region.to_excel(writer, sheet_name='Report by Region', index=False)
        report_category.to_excel(writer, sheet_name='Report by Category', index=False)
        
        # --- Helper to format a single report using openpyxl ---
        def apply_openpyxl_formatting(df, sheetname):
            worksheet = writer.sheets[sheetname]
            
            # Iterate through columns to apply width and header style
            for col_idx, column in enumerate(df.columns):
                # 1. Calculate column width
                header_len = len(column)
                # Find max length of all data in the column (only look at first 100 rows for efficiency)
                max_len = df[column].astype(str).str.len().head(100).max() if len(df) > 0 else 0
                adjusted_width = max(max(max_len or 0, header_len), 10) + 2 # Min width 10, add padding 2
                col_letter = get_column_letter(col_idx + 1)
                worksheet.column_dimensions[col_letter].width = adjusted_width
                
                # 2. Apply Header Styling (Row 1)
                header_cell = worksheet.cell(row=1, column=col_idx + 1)
                header_cell.font = header_font
                header_cell.border = thin_border
                header_cell.alignment = Alignment(horizontal='center')
                
                # 3. Apply Data Styling (Rows 2 onwards)
                for row_idx in range(2, len(df) + 2):
                    data_cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                    data_cell.border = thin_border
                    
                    # Apply currency format to relevant columns
                    if column in ['sales_amount', 'sales_cost', 'profit']:
                        data_cell.number_format = '#,##0.00' # Currency format with comma separator

        # Apply formatting to both sheets
        apply_openpyxl_formatting(report_region, 'Report by Region')
        apply_openpyxl_formatting(report_category, 'Report by Category')

    print(f"   - Export complete. Check '{filename}' for results.")


def main(region=None, category=None):
    """
    Main function to run the sales data analysis workflow.
    
    Arguments:
        region (str, optional): Filter by store region. Defaults to None (no filter).
        category (str, optional): Filter by product category. Defaults to None (no filter).
    """
    
    # Generate a descriptive filename based on the filters applied
    filter_desc = "ALL_DATA"
    if region and category:
        filter_desc = f"{region}_and_{category}"
    elif region:
        filter_desc = f"{region}_Region"
    elif category:
        filter_desc = f"{category}_Category"
        
    output_filename = f"sales_reports_{filter_desc}.xlsx"
    
    print(f"\n--- Running Analysis for: {filter_desc.replace('_', ' ')} ---")
    
    try:
        # 1. Data Retrieval (only happens once regardless of filters)
        sales_df, product_df, store_df, currency_df = load_data()
        
        # 2. Data Processing and Conversion
        processed_df = process_sales_data(
            sales_df, product_df, store_df, currency_df, 
            store_region_filter=region, product_category_filter=category
        )
        
        if processed_df.empty:
            print("\nNo data remaining after applying filters. Reports cannot be generated.")
            return

        # 3. Sales Aggregation
        report_region, report_category = aggregate_reports(processed_df)
        
        # 4. Excel Export
        export_to_excel(report_region, report_category, output_filename)

    except Exception as e:
        print(f"\nAn error occurred: {e}")

# --- Example Usage ---
if __name__ == '__main__':
    # 1. Get Region Filter
    region_input = input("Enter Store Region filter (e.g., North, South, East, West, Central) or press Enter for NO filter: ")
    # Convert empty string input to None
    region_filter = region_input.strip() if region_input.strip() else None

    # 2. Get Category Filter
    category_input = input("Enter Product Category filter (e.g., Electronics, Clothing, Groceries) or press Enter for NO filter: ")
    # Convert empty string input to None
    category_filter = category_input.strip() if category_input.strip() else None

    print("-" * 50)
    print(f"Running analysis with: Region='{region_filter}' | Category='{category_filter}'")
    print("-" * 50)

    # 3. Run the main function with the user-provided filters
    # The main function handles the logic for 0, 1, or 2 filters.
    main(region=region_filter, category=category_filter)